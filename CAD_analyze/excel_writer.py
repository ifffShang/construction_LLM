"""
Excel writer — produces an .xlsx file matching the format of
NHC_ADNOC_Architectural_Material_Tracker_by_Building.xlsx
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from templates import DEFAULT_PROJECT_METADATA

# ── colours ──────────────────────────────────────────────────────────────────
DARK_BLUE  = "1F4E78"
WHITE      = "FFFFFF"
LIGHT_BLUE = "D9EAF7"
DARK_GRAY  = "1F1F1F"
MID_GRAY   = "666666"
ALT_ROW    = "EBF3FB"
HOLD_ROW   = "FFFF00"
UNCONFIRMED = "FFF3CD"   # light amber for materials not confirmed by CAD callout


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _thin() -> Border:
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _med_bottom() -> Border:
    return Border(bottom=Side(style="medium"))


# ── Sheet 1 headers & widths ──────────────────────────────────────────────────
TRACKER_HEADERS = [
    ("SN",                      5),
    ("Building",               13),
    ("Unit",                    8),
    ("System",                 14),
    ("Material Category",      18),
    ("Material Name",          24),
    ("ADNOC Spec Ref",         22),
    ("International Standard", 18),
    ("Description",            32),
    ("Approved Vendor by ADNOC", 38),
    ("Brand",                  12),
    ("Origin",                 10),
    ("Unit",                    7),
    ("Quantity",               10),
    ("Drawing Ref",            22),
    ("Submittal Status",       16),
    ("Consultant Approval",    18),
    ("Client Approval",        16),
    ("AVL Status",             12),
    ("PO Status",              12),
    ("Production Status",      16),
    ("Delivery Status",        14),
    ("Site Status",            12),
    ("Installation Status",    16),
]

QTY_COL_INDEX = 14    # 1-based column index of "Quantity"
BUILDING_COL   = 2    # column B


def write_tracker(
    rows: list[dict],
    quantity_basis: list[dict],
    avl_reference: list[dict],
    output_path: str,
    project_metadata: dict | None = None,
):
    meta = {**DEFAULT_PROJECT_METADATA, **(project_metadata or {})}
    wb = openpyxl.Workbook()

    # ── Sheet 1 ───────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "ADNOC Material Tracker"

    # Row 1: title
    ws.cell(1, 1, "Architectural & Structural Material Tracker by Building").font = Font(
        bold=True, size=13, color=DARK_GRAY
    )
    ws.cell(1, 2, "Architectural and structural materials extracted from CAD drawings and specification documents").font = Font(
        size=10, color=MID_GRAY
    )
    ws.row_dimensions[1].height = 22

    # Rows 2-7: metadata
    meta_rows = [
        ("Project Name",    meta.get("project_name", "")),
        ("Owner",           meta.get("owner", "")),
        ("EPCM",            meta.get("epcm", "")),
        ("Contractor",      meta.get("contractor", "")),
        ("Sub-Contractor",  meta.get("sub_contractor", "")),
        ("Project No.",     meta.get("project_no", "")),
    ]
    for i, (label, value) in enumerate(meta_rows, start=2):
        lc = ws.cell(i, 1, label)
        lc.font = Font(bold=True, size=10)
        lc.fill = _fill(LIGHT_BLUE)
        ws.cell(i, 2, value).font = Font(size=10)
        ws.row_dimensions[i].height = 16

    # Row 8: blank spacer
    ws.row_dimensions[8].height = 6

    # Row 9: headers
    for col_idx, (header, width) in enumerate(TRACKER_HEADERS, start=1):
        cell = ws.cell(9, col_idx, header)
        cell.fill      = _fill(DARK_BLUE)
        cell.font      = Font(bold=True, size=10, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _med_bottom()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[9].height = 36

    # Data rows starting at row 10
    qty_col_letter = get_column_letter(QTY_COL_INDEX)

    for i, row in enumerate(rows):
        r = 10 + i

        is_hold = row.get("submittal_status", "").upper() == "HOLD"
        not_confirmed = not row.get("confirmed_in_cad", True)

        if is_hold:
            row_fill = _fill(HOLD_ROW)
        elif not_confirmed:
            row_fill = _fill(UNCONFIRMED)
        elif i % 2 == 1:
            row_fill = _fill(ALT_ROW)
        else:
            row_fill = None

        values = [
            row["sn"],
            row["building"],
            row.get("unit", ""),
            row["system"],
            row["material_category"],
            row["material_name"],
            row["adnoc_spec_ref"],
            row["international_standard"],
            row["description"],
            row["approved_vendor"],
            row["brand"],
            row["origin"],
            row["unit"],
            None,                          # Quantity — formula or value
            row["drawing_ref"],
            row["submittal_status"],
            row["consultant_approval"],
            row["client_approval"],
            row["avl_status"],
            row["po_status"],
            row["production_status"],
            row["delivery_status"],
            row["site_status"],
            row["installation_status"],
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(r, col_idx, val)
            cell.font      = Font(size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border    = _thin()
            if row_fill:
                cell.fill = row_fill

        # Quantity cell: VLOOKUP for m² roof materials, else plain value
        qty_cell = ws.cell(r, QTY_COL_INDEX)
        qty_raw = row.get("quantity", "TBD")
        if row["unit"] == "m²" and row["system"] == "Roof":
            qty_cell.value = (
                f'=IFERROR(VLOOKUP({get_column_letter(BUILDING_COL)}{r},'
                f"'Quantity Basis'!A:B,2,FALSE),\"\")"
            )
        else:
            # Try numeric; fall back to string
            try:
                qty_cell.value = float(qty_raw)
            except (ValueError, TypeError):
                qty_cell.value = qty_raw
        qty_cell.font      = Font(size=9)
        qty_cell.alignment = Alignment(vertical="center")
        qty_cell.border    = _thin()
        if row_fill:
            qty_cell.fill = row_fill

        ws.row_dimensions[r].height = 28

    # Freeze pane: keep header + building/SN visible while scrolling
    ws.freeze_panes = "D10"

    # ── Sheet 2: Quantity Basis ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Quantity Basis")
    qb_headers = [
        ("Building", 14),
        ("Approx Roof Area (m²)", 22),
        ("Basis / Note", 55),
    ]
    for col_idx, (h, w) in enumerate(qb_headers, start=1):
        cell = ws2.cell(1, col_idx, h)
        cell.fill      = _fill(DARK_BLUE)
        cell.font      = Font(bold=True, size=10, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws2.column_dimensions[get_column_letter(col_idx)].width = w
    ws2.row_dimensions[1].height = 28

    for i, qb in enumerate(quantity_basis):
        r = i + 2
        values = [qb["building"], qb["roof_area_m2"], qb["note"]]
        for col_idx, val in enumerate(values, start=1):
            cell = ws2.cell(r, col_idx, val)
            cell.font      = Font(size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border    = _thin()
            if i % 2 == 1:
                cell.fill = _fill(ALT_ROW)
        ws2.row_dimensions[r].height = 20

    # ── Sheet 3: AVL Reference ────────────────────────────────────────────────
    ws3 = wb.create_sheet("AVL Reference")
    avl_headers = [
        ("Material Name", 30),
        ("ADNOC-Approved Suppliers / Manufacturers", 60),
        ("Notes / Performance Requirements", 45),
    ]
    for col_idx, (h, w) in enumerate(avl_headers, start=1):
        cell = ws3.cell(1, col_idx, h)
        cell.fill      = _fill(DARK_BLUE)
        cell.font      = Font(bold=True, size=10, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws3.column_dimensions[get_column_letter(col_idx)].width = w
    ws3.row_dimensions[1].height = 28

    for i, avl in enumerate(avl_reference):
        r = i + 2
        values = [avl["material_name"], avl["vendors"], avl.get("notes", "")]
        for col_idx, val in enumerate(values, start=1):
            cell = ws3.cell(r, col_idx, val)
            cell.font      = Font(size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border    = _thin()
            if i % 2 == 1:
                cell.fill = _fill(ALT_ROW)
        ws3.row_dimensions[r].height = 22

    wb.save(output_path)
    return output_path
